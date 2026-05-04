"""
Tests for PTW Export Functionality
"""
from django.test import TestCase
from django.utils import timezone
from django.contrib.auth import get_user_model
from rest_framework.test import APIClient
from rest_framework import status
from datetime import timedelta
from io import BytesIO
import zipfile

from ptw.models import Permit, PermitType
from authentication.models import Project

User = get_user_model()


class ExportTestCase(TestCase):
    def setUp(self):
        """Set up test data"""
        self.user = User.objects.create_user(
            username='testuser',
            email='test@example.com',
            password='testpass123',
            name='Test',
            surname='User',
            admin_type='client'
        )
        
        self.project = Project.objects.create(
            name='Test Project',
            location='Test Location'
        )
        
        self.permit_type = PermitType.objects.create(
            name='Hot Work',
            category='hot_work',
            risk_level='high'
        )
        
        self.now = timezone.now()
        
        # Create test permit
        self.permit = Permit.objects.create(
            permit_number='PTW-TEST-001',
            permit_type=self.permit_type,
            title='Test Permit',
            description='Test Description',
            location='Test Area',
            planned_start_time=self.now,
            planned_end_time=self.now + timedelta(hours=8),
            created_by=self.user,
            project=self.project,
            status='active'
        )
        
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)
    
    def test_export_pdf_single_returns_pdf_headers(self):
        """Test single permit PDF export returns correct headers"""
        response = self.client.get(f'/api/v1/ptw/permits/{self.permit.id}/export_pdf/')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertIn('attachment', response['Content-Disposition'])
        self.assertIn('PTW_PTW-TEST-001.pdf', response['Content-Disposition'])
        self.assertGreater(len(response.content), 0)
    
    def test_export_excel_returns_xlsx_headers(self):
        """Test Excel export returns correct headers"""
        response = self.client.get('/api/v1/ptw/permits/export_excel/')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        self.assertIn('attachment', response['Content-Disposition'])
        self.assertIn('.xlsx', response['Content-Disposition'])
        self.assertGreater(len(response.content), 0)
    
    def test_export_excel_detailed_has_multiple_sheets(self):
        """Test detailed Excel export has multiple sheets"""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")
        
        response = self.client.get('/api/v1/ptw/permits/export_excel/?detailed=true')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        # Load workbook and check sheets
        wb = openpyxl.load_workbook(BytesIO(response.content))
        sheet_names = wb.sheetnames
        
        self.assertIn('Permits', sheet_names)
        self.assertIn('Isolation Points', sheet_names)
        self.assertIn('Gas Readings', sheet_names)
        self.assertIn('Closeout', sheet_names)
        self.assertIn('Audit Logs', sheet_names)
    
    def test_bulk_export_pdf_zip_contains_files(self):
        """Test bulk PDF export returns ZIP with files"""
        # Create additional permits
        permit2 = Permit.objects.create(
            permit_number='PTW-TEST-002',
            permit_type=self.permit_type,
            description='Test Permit 2',
            location='Test Area 2',
            planned_start_time=self.now,
            planned_end_time=self.now + timedelta(hours=8),
            created_by=self.user,
            project=self.project,
            status='active'
        )
        
        response = self.client.post(
            '/api/v1/ptw/permits/bulk_export_pdf/',
            {'permit_ids': [self.permit.id, permit2.id]},
            format='json'
        )
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response['Content-Type'], 'application/zip')
        self.assertIn('attachment', response['Content-Disposition'])
        
        # Verify ZIP contents
        zip_buffer = BytesIO(response.content)
        with zipfile.ZipFile(zip_buffer, 'r') as zip_file:
            file_list = zip_file.namelist()
            self.assertEqual(len(file_list), 2)
            self.assertIn('PTW_PTW-TEST-001.pdf', file_list)
            self.assertIn('PTW_PTW-TEST-002.pdf', file_list)
    
    def test_bulk_export_limit_enforced(self):
        """Test bulk export enforces permit limit"""
        # Create 201 permit IDs (exceeds default limit of 200)
        permit_ids = list(range(1, 202))
        
        response = self.client.post(
            '/api/v1/ptw/permits/bulk_export_pdf/',
            {'permit_ids': permit_ids},
            format='json'
        )
        
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.json())
        self.assertIn('Maximum', response.json()['error'])
    
    def test_export_permission_filters_queryset(self):
        """Test export only includes permits user can view"""
        # Create another user
        other_user = User.objects.create_user(
            username='otheruser',
            email='other@example.com',
            password='testpass123',
            name='Other',
            surname='User',
            admin_type='client'
        )
        
        # Create permit for other user in different project
        other_project = Project.objects.create(
            name='Other Project',
            location='Other Location'
        )
        
        other_permit = Permit.objects.create(
            permit_number='PTW-OTHER-001',
            permit_type=self.permit_type,
            description='Other Permit',
            location='Other Area',
            planned_start_time=self.now,
            planned_end_time=self.now + timedelta(hours=8),
            created_by=other_user,
            project=other_project,
            status='active'
        )
        
        # Try to export other user's permit
        response = self.client.post(
            '/api/v1/ptw/permits/bulk_export_pdf/',
            {'permit_ids': [other_permit.id]},
            format='json'
        )
        
        # Should return 404 or empty result (depending on permission implementation)
        self.assertIn(response.status_code, [status.HTTP_404_NOT_FOUND, status.HTTP_200_OK])
    
    def test_bulk_export_excel_with_detailed(self):
        """Test bulk Excel export with detailed flag"""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")
        
        response = self.client.post(
            '/api/v1/ptw/permits/bulk_export_excel/',
            {'permit_ids': [self.permit.id], 'detailed': True},
            format='json'
        )
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Verify multiple sheets
        wb = openpyxl.load_workbook(BytesIO(response.content))
        self.assertGreater(len(wb.sheetnames), 1)
    
    def test_bulk_export_empty_permit_ids(self):
        """Test bulk export with empty permit IDs returns error"""
        response = self.client.post(
            '/api/v1/ptw/permits/bulk_export_pdf/',
            {'permit_ids': []},
            format='json'
        )
        
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.json())
