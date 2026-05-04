"""
PTW Excel Export Utilities
Multi-sheet Excel export with formatting
"""
from datetime import datetime
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def generate_excel_export(queryset, detailed=False):
    """
    Generate Excel workbook for permits
    
    Args:
        queryset: Permit queryset
        detailed: If True, include additional sheets
    
    Returns:
        openpyxl.Workbook
    """
    wb = openpyxl.Workbook()
    
    # Main permits sheet
    _generate_permits_sheet(wb, queryset)
    
    if detailed:
        # Additional detail sheets
        _generate_isolation_sheet(wb, queryset)
        _generate_gas_readings_sheet(wb, queryset)
        _generate_closeout_sheet(wb, queryset)
        _generate_audit_logs_sheet(wb, queryset)
    
    return wb


def _generate_permits_sheet(wb, queryset):
    """Generate main permits sheet"""
    ws = wb.active
    ws.title = "Permits"
    
    # Headers
    headers = [
        'Permit Number', 'Title', 'Type', 'Category', 'Status', 'Priority',
        'Risk Level', 'Risk Score', 'Location', 'Project',
        'Planned Start', 'Planned End', 'Actual Start', 'Actual End',
        'Created By', 'Issuer', 'Receiver', 'Verifier', 'Approver',
        'Created At', 'Submitted At', 'Verified At', 'Approved At'
    ]
    
    # Write headers with formatting
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1890FF", end_color="1890FF", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Freeze top row
    ws.freeze_panes = 'A2'
    
    # Data rows
    for row_idx, permit in enumerate(queryset, 2):
        data = [
            permit.permit_number,
            permit.title or permit.description[:50],
            permit.permit_type.name,
            permit.permit_type.get_category_display(),
            permit.get_status_display(),
            permit.get_priority_display(),
            permit.get_risk_level_display(),
            permit.risk_score,
            permit.location,
            permit.project.name if permit.project else 'N/A',
            permit.planned_start_time,
            permit.planned_end_time,
            permit.actual_start_time,
            permit.actual_end_time,
            f"{permit.created_by.name} {permit.created_by.surname}".strip() if permit.created_by else 'N/A',
            f"{permit.issuer.name} {permit.issuer.surname}".strip() if permit.issuer else 'N/A',
            f"{permit.receiver.name} {permit.receiver.surname}".strip() if permit.receiver else 'N/A',
            f"{permit.verifier.name} {permit.verifier.surname}".strip() if permit.verifier else 'N/A',
            f"{permit.approved_by.name} {permit.approved_by.surname}".strip() if permit.approved_by else 'N/A',
            permit.created_at,
            permit.submitted_at,
            permit.verified_at,
            permit.approved_at
        ]
        
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Format dates
            if isinstance(value, datetime):
                cell.number_format = 'YYYY-MM-DD HH:MM:SS'
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Enable filters
    ws.auto_filter.ref = ws.dimensions


def _generate_isolation_sheet(wb, queryset):
    """Generate isolation points sheet"""
    ws = wb.create_sheet("Isolation Points")
    
    headers = [
        'Permit Number', 'Point Code/Name', 'Point Type', 'Energy Type',
        'Lock Applied', 'Lock Count', 'Lock IDs', 'Status',
        'Isolated By', 'Isolated At', 'Verified By', 'Verified At',
        'Deisolated By', 'Deisolated At'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="52C41A", end_color="52C41A", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.freeze_panes = 'A2'
    
    # Data
    row_idx = 2
    for permit in queryset:
        for point in permit.isolation_points.all():
            point_name = point.point.point_code if point.point else point.custom_point_name
            lock_ids = ', '.join(point.lock_ids) if point.lock_ids else 'N/A'
            
            data = [
                permit.permit_number,
                point_name,
                point.point.get_point_type_display() if point.point else 'Custom',
                point.point.get_energy_type_display() if point.point else 'N/A',
                'Yes' if point.lock_applied else 'No',
                point.lock_count,
                lock_ids,
                point.get_status_display(),
                point.isolated_by.username if point.isolated_by else 'N/A',
                point.isolated_at,
                point.verified_by.username if point.verified_by else 'N/A',
                point.verified_at,
                point.deisolated_by.username if point.deisolated_by else 'N/A',
                point.deisolated_at
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if isinstance(value, datetime):
                    cell.number_format = 'YYYY-MM-DD HH:MM:SS'
            
            row_idx += 1
    
    # Auto-adjust columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    ws.auto_filter.ref = ws.dimensions


def _generate_gas_readings_sheet(wb, queryset):
    """Generate gas readings sheet"""
    ws = wb.create_sheet("Gas Readings")
    
    headers = [
        'Permit Number', 'Gas Type', 'Reading', 'Unit', 'Acceptable Range',
        'Status', 'Tested By', 'Tested At', 'Equipment Used', 'Calibration Date'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FAAD14", end_color="FAAD14", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.freeze_panes = 'A2'
    
    # Data
    row_idx = 2
    for permit in queryset:
        for reading in permit.gas_readings.all():
            data = [
                permit.permit_number,
                reading.get_gas_type_display(),
                reading.reading,
                reading.unit,
                reading.acceptable_range,
                reading.status.upper(),
                reading.tested_by.username if reading.tested_by else 'N/A',
                reading.tested_at,
                reading.equipment_used or 'N/A',
                reading.calibration_date
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if isinstance(value, datetime):
                    cell.number_format = 'YYYY-MM-DD HH:MM:SS'
            
            row_idx += 1
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    ws.auto_filter.ref = ws.dimensions


def _generate_closeout_sheet(wb, queryset):
    """Generate closeout checklist sheet"""
    ws = wb.create_sheet("Closeout")
    
    headers = [
        'Permit Number', 'Template Name', 'Item', 'Required', 'Done',
        'Completed By', 'Completed At', 'Closeout Completed', 'Closeout Completed By', 'Closeout Completed At'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="722ED1", end_color="722ED1", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.freeze_panes = 'A2'
    
    # Data
    row_idx = 2
    for permit in queryset:
        try:
            closeout = permit.closeout
            if closeout and closeout.template:
                for item in closeout.template.items:
                    key = item['key']
                    status_info = closeout.checklist.get(key, {})
                    
                    data = [
                        permit.permit_number,
                        closeout.template.name,
                        item['label'],
                        'Yes' if item.get('required', False) else 'No',
                        'Yes' if status_info.get('done', False) else 'No',
                        status_info.get('by', 'N/A'),
                        status_info.get('at', 'N/A'),
                        'Yes' if closeout.completed else 'No',
                        closeout.completed_by.username if closeout.completed_by else 'N/A',
                        closeout.completed_at
                    ]
                    
                    for col_idx, value in enumerate(data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        if isinstance(value, datetime):
                            cell.number_format = 'YYYY-MM-DD HH:MM:SS'
                    
                    row_idx += 1
        except Exception:
            continue
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    ws.auto_filter.ref = ws.dimensions


def _generate_audit_logs_sheet(wb, queryset):
    """Generate audit logs sheet"""
    ws = wb.create_sheet("Audit Logs")
    
    headers = [
        'Permit Number', 'Action', 'User', 'Timestamp', 'Comments', 'IP Address'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FF4D4F", end_color="FF4D4F", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.freeze_panes = 'A2'
    
    # Data (limit to recent logs to avoid huge files)
    row_idx = 2
    for permit in queryset:
        for log in permit.audit_logs.all().order_by('-timestamp')[:50]:  # Limit per permit
            data = [
                permit.permit_number,
                log.get_action_display(),
                log.user.username if log.user else 'System',
                log.timestamp,
                log.comments or '',
                log.ip_address or 'N/A'
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if isinstance(value, datetime):
                    cell.number_format = 'YYYY-MM-DD HH:MM:SS'
            
            row_idx += 1
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    ws.auto_filter.ref = ws.dimensions
